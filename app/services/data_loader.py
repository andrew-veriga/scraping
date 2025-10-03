import pandas as pd
import logging
from typing import Dict, Any, List, Tuple
from datetime import timezone
import os

def is_admin(message_ID):
    if message_ID in [
        '862550907349893151',
        '466815633347313664',
        '997105563123064892',
        '457962750644060170'
    ]:
        return True
    return False
def load_and_preprocess_data(file_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Loads and preprocesses Discord data from Excel/CSV file.
    Uses openpyxl for XLSX files and robust encoding detection for CSV files.
    Handles Chinese characters, emojis, and problematic encodings.
    """
    logging.info(f"Loading and preprocessing data from {file_path}")
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    messages_df = None
    
    if file_path.endswith('.xlsx'):
        # Use openpyxl for XLSX files
        try:
            logging.info("Loading XLSX file using openpyxl...")
            from openpyxl import load_workbook
            
            # Load workbook
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # Get all data from the worksheet
            data = []
            headers = None
            
            for row in worksheet.iter_rows(values_only=True):
                if headers is None:
                    # First row contains headers
                    headers = [str(cell) if cell is not None else '' for cell in row]
                    logging.info(f"Found headers: {headers}")
                else:
                    # Data rows
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    data.append(row_data)
            
            workbook.close()
            
            # Convert to DataFrame
            if headers and data:
                messages_df = pd.DataFrame(data, columns=headers)
                logging.info(f"✅ Successfully loaded XLSX file using openpyxl: {len(messages_df)} rows")
            else:
                raise ValueError("No data found in XLSX file")
                
        except Exception as e:
            logging.error(f"Failed to load XLSX with openpyxl: {e}")
            # Fallback to pandas
            
    elif file_path.endswith('.csv'):
        # Use robust encoding detection for CSV files
        encodings_to_try = [
            'utf-8-sig',    # UTF-8 with BOM
            'utf-8',        # Standard UTF-8
            'gbk',          # Chinese Windows encoding
            'gb2312',       # Simplified Chinese
            'big5',         # Traditional Chinese
            'utf-16le',     # UTF-16 Little Endian
            'utf-16',       # UTF-16 with BOM
            'latin-1',      # Fallback encoding
            'cp1252'        # Windows-1252
        ]
        
        for encoding in encodings_to_try:
            try:
                logging.info(f"Trying to load CSV with encoding: {encoding}")
                messages_df = pd.read_csv(
                    filepath_or_buffer=file_path,
                    dtype={
                        'Referenced Message ID': str, 
                        'Message ID': str, 
                        'Author ID': str,
                        'Content': str
                    },
                    encoding=encoding,
                    encoding_errors='replace',  # Replace problematic characters
                    engine='python',
                    na_values=['', 'nan', ' ', '\u200b', '\u200c', '\u200d'],
                    keep_default_na=True,
                )
                logging.info(f"✅ Successfully loaded CSV file with encoding: {encoding}")
                break
            except UnicodeDecodeError as e:
                logging.warning(f"UnicodeDecodeError with {encoding}: {e}")
                continue
            except UnicodeError as e:
                logging.warning(f"UnicodeError with {encoding}: {e}")
                continue
            except Exception as e:
                logging.warning(f"Other error with {encoding}: {e}")
                continue
        
        if messages_df is None:
            # Nuclear option: binary read and clean
            logging.warning("All standard encodings failed. Trying binary cleaning...")
            try:
                with open(file_path, 'rb') as f:
                    content = f.read()
                    # Remove common problematic bytes
                    content = content.replace(b'\x00', b'')  # Remove null bytes
                    content = content.replace(b'\xff\xfe', b'')  # Remove UTF-16 BOM
                    content = content.replace(b'\xfe\xff', b'')  # Remove UTF-16 BE BOM
                    content = content.replace(b'\xef\xbb\xbf', b'')  # Remove UTF-8 BOM
                    
                    # Try to decode
                    try:
                        decoded_content = content.decode('utf-8', errors='replace')
                    except:
                        decoded_content = content.decode('latin-1', errors='replace')
                    
                    from io import StringIO
                    messages_df = pd.read_csv(
                        StringIO(decoded_content),
                        dtype={
                            'Referenced Message ID': str, 
                            'Message ID': str, 
                            'Author ID': str,
                            'Content': str
                        },
                        engine='python',
                        na_values=['', 'nan', ' ', '\u200b', '\u200c', '\u200d'],
                        keep_default_na=True,
                    )
                    logging.info("✅ Successfully loaded CSV file with binary cleaning")
            except Exception as e:
                logging.error(f"Binary cleaning also failed: {e}")
                raise ValueError(f"Could not load CSV file with any method: {e}")
    
    else:
        raise ValueError(f"Unsupported file format: {file_path}")
    
    if messages_df is None:
        raise ValueError(f"Could not load file {file_path}")
    
    logging.info(f"File loaded successfully: {len(messages_df)} rows with columns: {list(messages_df.columns)}")
    
    # Remove leading apostrophes from ID fields
    id_columns = ['Message ID', 'Author ID', 'Referenced Message ID']
    for col in id_columns:
        if col in messages_df.columns:
            # Remove leading apostrophes from ID fields
            messages_df[col] = messages_df[col].astype(str).str.lstrip("'")
            logging.info(f"Removed leading apostrophes from {col} column")
    
    # Clean zero-width spaces from all string columns
    unicode_chars_to_remove = ['\u200b', '\u200c', '\u200d', '\ufeff']  # Zero-width spaces and BOM
    for col in messages_df.columns:
        if messages_df[col].dtype == 'object':  # String columns
            for char in unicode_chars_to_remove:
                messages_df[col] = messages_df[col].astype(str).str.replace(char, '', regex=False)

    # Clean and preprocess data
    messages_df['Referenced Message ID'] = messages_df['Referenced Message ID'].fillna('')
    messages_df['Message ID'] = messages_df['Message ID'].astype(str)
    messages_df['Author ID'] = messages_df['Author ID'].astype(str)
    
    # Convert timestamps
    if 'Unix Timestamp' in messages_df.columns:
        messages_df['DateTime'] = pd.to_datetime(messages_df['Unix Timestamp'], unit='s', utc=True)
        # Remove Unix Timestamp column (no longer needed)
        messages_df = messages_df.drop(columns=['Unix Timestamp'], errors='ignore')
    elif 'DateTime' in messages_df.columns:
        # If DateTime column already exists, ensure it's timezone-aware
        messages_df['DateTime'] = pd.to_datetime(messages_df['DateTime'])
        if messages_df['DateTime'].dt.tz is None:
            messages_df['DateTime'] = messages_df['DateTime'].dt.tz_localize('UTC')
        else:
            # If already timezone-aware, convert to UTC
            messages_df['DateTime'] = messages_df['DateTime'].dt.tz_convert('UTC')
    
    # Set Message ID as index for easy lookup
    messages_df.set_index('Message ID', inplace=True, drop=False)
    
    # Sort by timestamp for proper processing order
    messages_df = messages_df.sort_values('DateTime')
    
    logging.info(f"{len(messages_df)} messages loaded and preprocessed successfully")
    
    # Extract authors from messages
    authors_df = extract_authors_from_messages(messages_df)
    
    return messages_df, authors_df


def extract_authors_from_messages(messages_df: pd.DataFrame) -> pd.DataFrame:
    """
    Extract unique authors from messages DataFrame.
    """
    # Get unique authors
    authors_data = []
    seen_authors = set()
    
    for _, row in messages_df.iterrows():
        author_id = str(row['Author ID'])
        author_name = str(row.get('Author Name', ''))
        
        if author_id not in seen_authors and author_id != 'nan':
            authors_data.append({
                'author_id': author_id,
                'author_name': author_name,
                'author_type': 'user'  # Default type
            })
            seen_authors.add(author_id)
    
    authors_df = pd.DataFrame(authors_data)
    logging.info(f"Extracted {len(authors_df)} unique authors")
    
    return authors_df

def load_and_preprocess_data_old(file_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Loads and preprocesses Discord data from CSV file.
    Enhanced to handle hierarchical message relationships.
    """
    messages_df = pd.read_csv(
        filepath_or_buffer=file_path,
        dtype={
            'Referenced Message ID': str, 
            'Message ID': str, 
            'Author ID': str,
            'Content': str
        },
        encoding='utf-8',
        encoding_errors='replace',  # This will replace problematic characters with
        engine='python',
        na_values=['', 'nan', ' ', '\u200b', '\u200c', '\u200d'],
        keep_default_na=True,
    )

    logging.info(f"Loaded {len(messages_df)} messages")
        # Convert timestamps
    if 'Unix Timestamp' in messages_df.columns:
        messages_df['DateTime'] = pd.to_datetime(messages_df['Unix Timestamp'], unit='s', utc=True)
        # Remove Unix Timestamp column (no longer needed)
        messages_df = messages_df.drop(columns=['Unix Timestamp'], errors='ignore')
    elif 'DateTime' in messages_df.columns:
        # If DateTime column already exists, ensure it's timezone-aware
        messages_df['DateTime'] = pd.Timestamp(messages_df['DateTime'])
        if messages_df['DateTime'].dt.tz is None:
            messages_df['DateTime'] = messages_df['DateTime'].dt.tz_localize('UTC')
        else:
            # If already timezone-aware, convert to UTC
            messages_df['DateTime'] = messages_df['DateTime'].dt.tz_convert('UTC')
    # Set Message ID as index for easy lookup
    messages_df.set_index('Message ID', inplace=True, drop=False)
    
    # Sort by timestamp for proper processing order
    messages_df = messages_df.sort_values('DateTime')
    
    logging.info(f"{len(messages_df)} messages loaded and preprocessed successfully")
    authors_df = extract_authors_from_messages(messages_df)
    

    return messages_df, authors_df


def load_messages_to_database(messages_df: pd.DataFrame) -> Dict[str, int]:
    """
    Load messages from DataFrame into database with enhanced statistics.
    Returns detailed statistics about the loading process.
    """
    try:
        from app.services.database import get_database_service
        
        db_service = get_database_service()
        stats = {
            'total_messages': len(messages_df),
            'new_messages_created': 0,
            'existing_messages_skipped': 0,
            'threads_created': 0
        }
        
        logging.info(f"Loading {len(messages_df)} messages into database...")
        
        # Convert DataFrame to list of dictionaries for database insertion
        messages_data = []
        for _, row in messages_df.iterrows():
            # Prepare message data with proper null handling and hierarchical fields
            # Remove leading apostrophes from ID fields
            message_id = str(row['Message ID']).replace('\u200b', '').lstrip("'")
            author_id = str(row['Author ID']).replace('\u200b', '').lstrip("'")
            referenced_message_id = str(row['Referenced Message ID']).lstrip("'") if pd.notna(row['Referenced Message ID']) and row['Referenced Message ID'] else None
            
            message_data = {
                'message_id': message_id,
                'parent_id': str(row.get('parent_id','')) if pd.notna(row.get('parent_id')) and row.get('parent_id') else None,
                'author_id': author_id,
                'content': str(row['Content']),
                'datetime': row['DateTime'],
                'referenced_message_id': referenced_message_id,
                'attachments': str(row.get('Attachments', '')) if pd.notna(row.get('Attachments')) and row.get('Attachments') else None,
                'thread_id': str(row['thread_id']) if pd.notna(row.get('thread_id')) and row.get('thread_id') else None
            }
            messages_data.append(message_data)
        
        # Try hierarchical method first, fallback to basic method
        try:
            # Use hierarchical method if available
            created_count = db_service.bulk_create_messages_hierarchical(messages_data)
        except AttributeError:
            # Fallback to basic method if hierarchical method doesn't exist
            logging.info("Hierarchical method not available, using basic method")
            # Remove hierarchical fields for basic method
            basic_messages_data = []
            for msg in messages_data:
                basic_msg = {k: v for k, v in msg.items() if k not in ['parent_id', 'thread_id']}
                basic_messages_data.append(basic_msg)
            created_count = db_service.bulk_create_messages(basic_messages_data)
        
        stats['new_messages_created'] = created_count
        stats['existing_messages_skipped'] = stats['total_messages'] - created_count
        stats['threads_created'] = 0  # No threads created in raw loading phase
        
        logging.info("Message loading complete:")
        logging.info(f"  ✅ New messages created: {stats['new_messages_created']}")
        logging.info(f"  ⏭️  Existing messages skipped: {stats['existing_messages_skipped']}")
        
        return stats
        
    except Exception as e:
        logging.error(f"Failed to load messages to database: {e}")
        # Return stats with error indication
        return {
            'total_messages': len(messages_df) if 'messages_df' in locals() else 0,
            'new_messages_created': 0,
            'existing_messages_skipped': 0,
            'threads_created': 0,
            'error': str(e)
        }


def load_authors_to_database(authors_df: pd.DataFrame) -> Dict[str, int]:
    """
    Load authors from DataFrame into database with enhanced statistics.
    Returns detailed statistics about the loading process.
    """
    try:
        from app.services.database import get_database_service
        
        db_service = get_database_service()
        stats = {
            'total_authors': len(authors_df),
            'new_authors_created': 0,
            'existing_authors_skipped': 0
        }
        
        logging.info(f"Loading {len(authors_df)} authors into database...")
        
        # Validate and clean author data
        cleaned_authors_data = []
        for _, row in authors_df.iterrows():
            # Ensure required fields are present
            
            required_fields = ['author_id', 'author_name', 'author_type']
            if not all(field in authors_df.columns for field in required_fields):
                logging.warning(f"DataFrame missing required columns: {required_fields}")
                break
            
            # Clean author data
            cleaned_author = {
                'author_id': str(row['author_id']).strip().replace('\u200b', '').lstrip("'"),
                'author_name': str(row['author_name']).strip(),
                'author_type': str(row['author_type']).strip()
            }
            
            # Skip if author_id is empty
            if not cleaned_author['author_id']:
                logging.warning(f"Skipping author with empty author_id: {row.to_dict()}")
                continue
                
            cleaned_authors_data.append(cleaned_author)
        
        # Bulk create authors in database
        created_count = db_service.bulk_create_authors(cleaned_authors_data)
        
        stats['new_authors_created'] = created_count
        stats['existing_authors_skipped'] = stats['total_authors'] - created_count
        
        logging.info("Author loading complete:")
        logging.info(f"  ✅ New authors created: {stats['new_authors_created']}")
        logging.info(f"  ⏭️  Existing authors skipped: {stats['existing_authors_skipped']}")
        
        return stats
        
    except Exception as e:
        logging.error(f"Failed to load authors to database: {e}")
        # Return stats with error indication
        return {
            'total_authors': len(authors_df) if 'authors_df' in locals() else 0,
            'new_authors_created': 0,
            'existing_authors_skipped': 0,
            'error': str(e)
        }


def extract_authors_from_messages(messages_df: pd.DataFrame) -> pd.DataFrame:
    """
    Extract unique authors from messages DataFrame.
    Returns a DataFrame with author data ready for database insertion.
    """
    try:
        # Get unique authors from the messages
        authors_df = messages_df[['Author ID', 'Author Name']].drop_duplicates()
        
        # Create new DataFrame with required columns
        authors_data = []
        for _, row in authors_df.iterrows():
            author_id = str(row['Author ID']).strip().lstrip("'")
            author_name = str(row['Author Name']).strip()
            
            # Skip if author_id is empty
            if not author_id:
                continue
            
            # Determine author type
            author_type = 'admin' if is_admin(author_id) else 'user'
            
            authors_data.append({
                'author_id': author_id,
                'author_name': author_name,
                'author_type': author_type
            })
        
        # Convert to DataFrame
        result_df = pd.DataFrame(authors_data)
        
        logging.info(f"Extracted {len(result_df)} unique authors from messages")
        return result_df
        
    except Exception as e:
        logging.error(f"Failed to extract authors from messages: {e}")
        return pd.DataFrame(columns=['author_id', 'author_name', 'author_type'])
